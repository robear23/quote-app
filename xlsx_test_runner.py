"""
xlsx_test_runner.py — XLSX quote generation test suite.

Two modes:
  --mode scratch   (default) Tests DocumentFactory.generate_xlsx() with synthetic brand_dna —
                   no AI required, runs in seconds per template.
  --mode template  Creates blank XLSX templates, runs through the full AI pipeline
                   (extract_brand_dna_from_xlsx → build_xlsx_field_mapping →
                   generate_from_xlsx_template), then scores the output.

Scoring (40 pts total):
  | Dimension         | Points | What is checked                                         |
  |-------------------|--------|----------------------------------------------------------|
  | Line Items        |   10   | All expected line item descriptions present in output    |
  | Calculation Acc.  |   10   | Subtotal, tax, and grand total values are correct        |
  | Template Fidelity |   10   | Business name and customer name present                  |
  | Field Mapping     |   10   | Scratch: structure check. Template: mapping completeness |

Status: ≥28 → PASS · 14–27 → WARN · <14 → FAIL

Usage:
  python xlsx_test_runner.py                     # scratch, all 10
  python xlsx_test_runner.py --mode template     # template pipeline, all 10
  python xlsx_test_runner.py --ids 04 06         # scratch, templates 04 and 06
  python xlsx_test_runner.py --mode template --ids 02 05
"""

import os
import re
import sys
import json
import random
import shutil
import logging
import argparse
import traceback
from pathlib import Path
from datetime import date, datetime

import openpyxl
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent))
from document_factory import DocumentFactory, _sym
from config import settings

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("xlsx_test_runner")

TEMPLATES_DIR = Path("test_templates")
GENERATED_DIR = Path("test_generated")
RESULTS_DIR = Path("test_results")
for _d in (TEMPLATES_DIR, GENERATED_DIR, RESULTS_DIR):
    _d.mkdir(exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE CONFIGURATIONS  (mirrors test_runner.py)
# ─────────────────────────────────────────────────────────────────────────────

TEMPLATES = [
    {
        "id": "01_aura_design",
        "field_style": "brackets",
        "business_name": "Aura Design Studio",
        "business_address": "Studio 12, Neon Works, London, E1 6QL",
        "contact_line": "hello@auradesign.io  |  www.auradesign.io",
        "columns": ["Project Item", "Amount"],
        "sample_rows": [
            ["Brand Identity System", "£2,500.00"],
            ["Web UI Kit - 15 Screens", "£1,800.00"],
        ],
        "totals_rows": [("Total Investment", "£4,300.00")],
        "client_label": "Design Prepared For",
        "ref_label": "Project ID",
        "date_label": "Issuance Date",
        "primary_color": "000000",
        "currency": "GBP",
        "tax_rate": 0.0,
        "quote_data": {
            "customer_name": "Zenith Ventures",
            "customer_address": "88 Kingsway, London, WC2B 6AA",
            "line_items": [
                {"description": "Brand Identity System (Concept + Style Guide)", "quantity": 1, "unit_price": 2500.00},
                {"description": "Web UI Kit - 15 Responsive Screens", "quantity": 1, "unit_price": 1800.00},
            ],
        },
    },
    {
        "id": "02_global_steel",
        "field_style": "labels",
        "business_name": "Global Steel & Supply Corp",
        "business_address": "1000 Industrial Way, Pittsburgh, PA 15201",
        "contact_line": "sales@globalsteel.com  |  Phone: 412-555-0100",
        "vat_reg": "EIN: 12-3456789",
        "columns": ["Part No", "Description", "Qty", "Unit Price", "Subtotal"],
        "sample_rows": [
            ["ST-405", "Structural Steel Beam 20ft", "10", "$1,250.00", "$12,500.00"],
            ["RB-12", "Reinforcement Bars 1/2in", "500", "$4.50", "$2,250.00"],
        ],
        "totals_rows": [("Subtotal", "$14,750.00"), ("Sales Tax (7%)", "$1,032.50"), ("TOTAL AMOUNT", "$15,782.50")],
        "client_label": "Consignee",
        "ref_label": "PO REF NO",
        "date_label": "SHIP DATE",
        "primary_color": "D35400",
        "currency": "USD",
        "tax_rate": 7.0,
        "quote_data": {
            "customer_name": "Midwest Bridge Builders",
            "customer_address": "500 Construction Road, Chicago, IL 60601",
            "line_items": [
                {"description": "ST-405 Structural Steel Beam 20ft", "quantity": 10, "unit_price": 1250.00},
                {"description": "RB-12 Reinforcement Bars 1/2in", "quantity": 500, "unit_price": 4.50},
                {"description": "Heavy Duty Steel Coupling Pins", "quantity": 100, "unit_price": 12.00},
            ],
        },
    },
    {
        "id": "03_azure_estate",
        "field_style": "sample",
        "business_name": "The Azure Estate",
        "business_address": "Valley Road, Napa Valley, CA 94558",
        "contact_line": "weddings@azureestate.com  |  Tel: 707-555-8888",
        "columns": ["Service Category", "Description", "Investment"],
        "sample_rows": [
            ["Venue", "Exclusive use of Grand Ballroom", "$8,500.00"],
            ["Catering", "Premium 5-course dinner", "$12,000.00"],
        ],
        "totals_rows": [("Subtotal", "$20,500.00"), ("Service Charge (18%)", "$3,690.00"), ("GRAND TOTAL", "$24,190.00")],
        "client_label": "Honoring",
        "ref_label": "Reservation #",
        "date_label": "Event Date",
        "primary_color": "2E4053",
        "currency": "USD",
        "tax_rate": 18.0,
        "quote_data": {
            "customer_name": "The Miller-Ross Wedding",
            "customer_address": "123 Vineyard Lane, St. Helena, CA 94574",
            "line_items": [
                {"description": "Exclusive use of the Grand Ballroom & Terrace", "quantity": 1, "unit_price": 8500.00},
                {"description": "Premium 5-course dinner per guest", "quantity": 100, "unit_price": 120.00},
                {"description": "Bridal Suite Accommodation (2 nights)", "quantity": 1, "unit_price": 1500.00},
            ],
        },
    },
    {
        "id": "04_quick_handyman",
        "field_style": "brackets",
        "business_name": "Quick Fix Handyman Services",
        "business_address": "12 Maple Ave, Springfield",
        "contact_line": "Call Mike: 555-0123",
        "columns": ["Job Description", "Price"],
        "sample_rows": [
            ["Repair leaky faucet in kitchen", "$85.00"],
            ["Replace door handle on front door", "$45.00"],
        ],
        "totals_rows": [("TOTAL DUE", "$130.00")],
        "client_label": "Customer",
        "ref_label": "Job ID",
        "date_label": "Date",
        "primary_color": "2874A6",
        "currency": "USD",
        "tax_rate": 0.0,
        "quote_data": {
            "customer_name": "Alice Peterson",
            "customer_address": "456 Oak Lane, Springfield",
            "line_items": [
                {"description": "Repair leaky faucet in kitchen", "quantity": 1, "unit_price": 85.00},
                {"description": "Replace door handle on front door", "quantity": 1, "unit_price": 45.00},
                {"description": "Clean gutters (rear of house)", "quantity": 1, "unit_price": 120.00},
            ],
        },
    },
    {
        "id": "05_vet_clinic",
        "field_style": "labels",
        "business_name": "Happy Paws Veterinary Clinic",
        "business_address": "77 Bark Way, Melbourne VIC 3000",
        "contact_line": "vets@happypaws.com.au  |  (03) 9876 5432",
        "columns": ["Procedure / Item", "Qty", "Cost (inc GST)"],
        "sample_rows": [
            ["Annual Vaccination & Health Check", "1", "$110.00"],
            ["Dental Scale and Polish", "1", "$350.00"],
        ],
        "totals_rows": [("GST Included", "$41.82"), ("TOTAL PAYABLE", "$460.00")],
        "client_label": "Pet Owner",
        "ref_label": "Patient Record #",
        "date_label": "Visit Date",
        "primary_color": "148F77",
        "currency": "AUD",
        "tax_rate": 10.0,
        "quote_data": {
            "customer_name": "Sarah Miller (Pet: Luna)",
            "customer_address": "15 Wattle St, Richmond VIC 3121",
            "line_items": [
                {"description": "Annual Vaccination & Health Check", "quantity": 1, "unit_price": 110.00},
                {"description": "Dental Scale and Polish", "quantity": 1, "unit_price": 350.00},
                {"description": "Antibiotic Course (10 days)", "quantity": 1, "unit_price": 45.00},
            ],
        },
    },
    {
        "id": "06_solar_rebate",
        "field_style": "sample",
        "business_name": "Eco-Power Solar Solutions",
        "business_address": "Unit 5, Energy Park, Sydney NSW 2000",
        "contact_line": "info@ecopower.com.au  |  ABN: 99 888 777 666",
        "columns": ["System Component", "Qty", "Price"],
        "sample_rows": [
            ["6.6kW Tier 1 Solar Panel Array", "1", "$6,500.00"],
            ["5kW Hybrid Inverter", "1", "$2,200.00"],
        ],
        "totals_rows": [("Gross System Total", "$8,700.00"), ("NET AMOUNT PAYABLE", "$6,300.00")],
        "client_label": "Applicant Name",
        "ref_label": "System Quote Ref",
        "date_label": "Expiry Date",
        "primary_color": "F1C40F",
        "currency": "AUD",
        "tax_rate": 0.0,
        "quote_data": {
            "customer_name": "Thomas Jenkins",
            "customer_address": "82 Botany Rd, Alexandria NSW 2015",
            "line_items": [
                {"description": "6.6kW Tier 1 Solar Panel Array", "quantity": 1, "unit_price": 6500.00},
                {"description": "5kW Hybrid Inverter (Battery Ready)", "quantity": 1, "unit_price": 2200.00},
                {"description": "Smart Meter Installation", "quantity": 1, "unit_price": 450.00},
            ],
        },
    },
    {
        "id": "07_move_easy",
        "field_style": "labels",
        "business_name": "Move Easy Logistics",
        "business_address": "202 Cargo Way, Dallas, TX 75201",
        "contact_line": "ops@moveeasy.com  |  DOT: 1234567",
        "columns": ["Service Description", "Volume (cu.ft)", "Rate", "Total"],
        "sample_rows": [
            ["Residential Move - 3 Bedroom House", "1200", "$4.50", "$5,400.00"],
            ["Packing & Unpacking Services", "1", "flat", "$850.00"],
        ],
        "totals_rows": [("Fuel Surcharge (10%)", "$625.00"), ("ESTIMATED TOTAL", "$6,875.00")],
        "client_label": "Customer Name",
        "ref_label": "Booking Ref",
        "date_label": "Proposed Move Date",
        "primary_color": "C0392B",
        "currency": "USD",
        "tax_rate": 0.0,
        "quote_data": {
            "customer_name": "Robert Henderson",
            "customer_address": "77 Sunset Blvd, Dallas, TX 75204",
            "line_items": [
                {"description": "Residential Move - 3 Bedroom House", "quantity": 1200, "unit_price": 4.50},
                {"description": "Packing & Unpacking Services (Flat Rate)", "quantity": 1, "unit_price": 850.00},
                {"description": "Full Insurance Coverage (Up to $50k)", "quantity": 1, "unit_price": 250.00},
            ],
        },
    },
    {
        "id": "08_code_crafters",
        "field_style": "brackets",
        "business_name": "Code Crafters Software",
        "business_address": "Tech Plaza, Berlin, Germany",
        "contact_line": "billing@codecrafters.de  |  VAT: DE987654321",
        "columns": ["Software Service", "Seats", "Monthly Fee", "Yearly Total"],
        "sample_rows": [
            ["Enterprise License - CRM Suite", "25", "€45.00", "€13,500.00"],
            ["Premium Support Add-on", "1", "€200.00", "€2,400.00"],
        ],
        "totals_rows": [("Subtotal (Annual)", "€15,900.00"), ("VAT (19%)", "€3,021.00"), ("TOTAL COST", "€18,921.00")],
        "client_label": "Account Owner",
        "ref_label": "License ID",
        "date_label": "Renewal Date",
        "primary_color": "2471A3",
        "currency": "EUR",
        "tax_rate": 19.0,
        "quote_data": {
            "customer_name": "Innova Solutions GmbH",
            "customer_address": "Alte Strasse 12, 10115 Berlin",
            "line_items": [
                {"description": "Enterprise License - CRM Suite", "quantity": 25, "unit_price": 540.00},
                {"description": "Premium Support Add-on (Annual)", "quantity": 1, "unit_price": 2400.00},
            ],
        },
    },
    {
        "id": "09_gourmet_catering",
        "field_style": "sample",
        "business_name": "Gourmet Garden Catering",
        "business_address": "High Street, Oxford, OX1 4AH",
        "contact_line": "events@gourmetgarden.co.uk",
        "columns": ["Item Description", "Price Per Head", "Quantity", "Total"],
        "sample_rows": [
            ["Artisan Hors d'oeuvres Selection", "£12.50", "120", "£1,500.00"],
            ["Plated Main Course - Seasonal Menu", "£45.00", "120", "£5,400.00"],
        ],
        "totals_rows": [("Subtotal", "£6,900.00"), ("TOTAL QUOTE", "£7,750.00")],
        "client_label": "Event Organized For",
        "ref_label": "Event Reference",
        "date_label": "Booking Date",
        "primary_color": "1D8348",
        "currency": "GBP",
        "tax_rate": 0.0,
        "quote_data": {
            "customer_name": "Oxford Alumni Association",
            "customer_address": "Sheldonian Theatre, Oxford, OX1 3AZ",
            "line_items": [
                {"description": "Artisan Hors d'oeuvres Selection", "quantity": 120, "unit_price": 12.50},
                {"description": "Plated Main Course - Seasonal Menu", "quantity": 120, "unit_price": 45.00},
                {"description": "Dessert Buffet & Coffee Station", "quantity": 120, "unit_price": 8.50},
            ],
        },
    },
    {
        "id": "10_fit_pro",
        "field_style": "labels",
        "business_name": "FitPro Personal Training",
        "business_address": "Health Hub, Gym Street, Manchester",
        "contact_line": "coach@fitpro.com",
        "columns": ["Training Package", "Cost"],
        "sample_rows": [
            ["10 Session Kickstart Pack", "£450.00"],
            ["Nutritional Consultation", "£85.00"],
        ],
        "totals_rows": [("TOTAL", "£535.00")],
        "client_label": "Client",
        "ref_label": "Membership #",
        "date_label": "Start Date",
        "primary_color": "7B241C",
        "currency": "GBP",
        "tax_rate": 0.0,
        "quote_data": {
            "customer_name": "Mark Stevens",
            "customer_address": "12 Broadway, Manchester, M1 1AA",
            "line_items": [
                {"description": "10 Session Kickstart Pack", "quantity": 1, "unit_price": 450.00},
                {"description": "Nutritional Consultation", "quantity": 1, "unit_price": 85.00},
                {"description": "FitPro T-Shirt & Water Bottle", "quantity": 1, "unit_price": 0.00},
            ],
        },
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# BLANK XLSX TEMPLATE CREATION
# ─────────────────────────────────────────────────────────────────────────────

def _field_val(style: str, bracket_text: str, sample_text: str = "") -> str:
    if style == "labels":
        return ""
    if style == "sample":
        return sample_text
    return bracket_text


def _hex_fill(wb: Workbook, hex_color: str):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=hex_color.lstrip("#").upper())


def _bold_font(wb: Workbook, color: str = "000000", size: int = 11):
    from openpyxl.styles import Font
    return Font(bold=True, color=color.lstrip("#").upper(), size=size)


def create_blank_xlsx(t: dict) -> Workbook:
    """Creates a blank XLSX template workbook for the given config dict."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"

    primary = t["primary_color"].lstrip("#").upper()
    fs = t.get("field_style", "brackets")

    white = "FFFFFF"
    light_grey = "F4F6F7"

    bold_white = Font(bold=True, color=white, size=12)
    bold_primary = Font(bold=True, color=primary, size=11)
    bold_std = Font(bold=True, size=10)
    std = Font(size=10)
    small_grey_font = Font(size=8, color="666666")

    fill_primary = PatternFill("solid", fgColor=primary)
    fill_light = PatternFill("solid", fgColor=light_grey)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right")

    row = 1

    # ── Business header ──────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    hdr_cell = ws.cell(row=row, column=1, value=t["business_name"])
    hdr_cell.font = bold_white
    hdr_cell.fill = fill_primary
    hdr_cell.alignment = center
    ws.row_dimensions[row].height = 22
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    addr_cell = ws.cell(row=row, column=1, value=f"{t['business_address']} | {t['contact_line']}")
    addr_cell.font = small_grey_font
    row += 1

    if t.get("vat_reg"):
        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row=row, column=1, value=t["vat_reg"]).font = small_grey_font
        row += 1

    row += 1  # spacer

    # ── Meta: ref + date ─────────────────────────────────────────────────────
    for (label, val_text, col_start) in [
        (t["ref_label"], _field_val(fs, f"[{t['ref_label']}]", "QT-1001"), 1),
        (t["date_label"], _field_val(fs, f"[{t['date_label']}]", "15 Jan 2024"), 3),
    ]:
        lc = ws.cell(row=row, column=col_start, value=label)
        lc.font = bold_std
        lc.fill = fill_light
        lc.border = thin_border
        vc = ws.cell(row=row, column=col_start + 1, value=val_text)
        vc.font = std
        vc.border = thin_border
    row += 1

    # Valid Until row
    ws.cell(row=row, column=1, value="Valid Until").font = bold_std
    ws.cell(row=row, column=1).fill = fill_light
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=_field_val(fs, "[EXPIRY_DATE]", "14 Feb 2024")).border = thin_border
    ws.cell(row=row, column=3, value="Status").font = bold_std
    ws.cell(row=row, column=3).fill = fill_light
    ws.cell(row=row, column=3).border = thin_border
    ws.cell(row=row, column=4, value=_field_val(fs, "[QUOTE_STATUS]", "Draft")).border = thin_border
    row += 2  # spacer

    # ── Bill To ──────────────────────────────────────────────────────────────
    bt = ws.cell(row=row, column=1, value="BILL TO:")
    bt.font = bold_primary
    row += 1

    ws.cell(row=row, column=1, value=_field_val(fs, f"[{t['client_label']}]", "ABC Company Ltd")).font = std
    ws.cell(row=row, column=2, value="Address:").font = bold_std
    ws.cell(row=row, column=3, value=_field_val(fs, "[Client Address]", "100 Example Street, City")).font = std
    row += 2  # spacer

    # ── Line items header ─────────────────────────────────────────────────────
    li_header_row = row
    columns = t["columns"]
    for ci, col_name in enumerate(columns, 1):
        hc = ws.cell(row=row, column=ci, value=col_name)
        hc.font = bold_white
        hc.fill = fill_primary
        hc.border = thin_border
        hc.alignment = center
    row += 1

    # ── Line items data rows ─────────────────────────────────────────────────
    for sample_row in t["sample_rows"]:
        for ci, val in enumerate(sample_row, 1):
            dc = ws.cell(row=row, column=ci, value=val)
            dc.font = std
            dc.border = thin_border
        row += 1
    row += 1  # spacer after line items

    # ── Totals ───────────────────────────────────────────────────────────────
    num_cols = len(columns)
    label_col = max(1, num_cols - 1)
    val_col = num_cols

    for ri, (label, val) in enumerate(t["totals_rows"]):
        is_last = ri == len(t["totals_rows"]) - 1
        lc = ws.cell(row=row, column=label_col, value=label)
        vc = ws.cell(row=row, column=val_col, value=val)
        lc.border = thin_border
        vc.border = thin_border
        lc.alignment = right_align
        vc.alignment = right_align
        if is_last:
            lc.font = bold_white
            vc.font = bold_white
            lc.fill = fill_primary
            vc.fill = fill_primary
        else:
            lc.font = bold_std
            vc.font = std
        row += 1

    row += 2
    # ── Footer ───────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    footer_cell = ws.cell(row=row, column=1, value=f"Please review the above quote. Valid for 14 days. | {t['business_name']}")
    footer_cell.font = small_grey_font

    # Set column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    return wb


# ─────────────────────────────────────────────────────────────────────────────
# CONTENT EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def extract_xlsx_text(path: Path) -> str:
    """Extract all non-empty cell values from an XLSX file as a single string."""
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        parts = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if val is not None and str(val).strip():
                        parts.append(str(val).strip())
        return "\n".join(parts)
    except Exception as e:
        return f"[TEXT EXTRACTION FAILED: {e}]"


def extract_xlsx_numbers(path: Path) -> list[float]:
    """Extract all numeric cell values from an XLSX file."""
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        nums = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        nums.append(float(cell.value))
        return nums
    except Exception:
        return []


# ─────────────────────────────────────────────────────────────────────────────
# SCORING
# ─────────────────────────────────────────────────────────────────────────────

def score_line_items(text: str, quote_data: dict) -> tuple[int, list[str]]:
    issues = []
    line_items = quote_data.get("line_items", [])
    found = 0
    for item in line_items:
        desc = item.get("description", "")
        needle = desc[:20].lower()
        if needle and needle in text.lower():
            found += 1
        else:
            issues.append(f"Description not found: '{desc[:40]}'")
    if not line_items:
        return 5, ["No line items in test data"]
    score = round((found / len(line_items)) * 10)
    if found == len(line_items):
        issues = []
    return score, issues


def score_calculations(text: str, numbers: list[float], quote_data: dict, tax_rate: float, sym: str) -> tuple[int, list[str]]:
    issues = []
    items = quote_data.get("line_items", [])
    subtotal = sum(float(i.get("quantity", 1)) * float(i.get("unit_price", 0)) for i in items)
    tax_amount = subtotal * (tax_rate / 100)
    grand_total = subtotal + tax_amount

    def _fmt(amount: float) -> str:
        return f"{sym}{amount:,.2f}"

    found = 0
    checks = [(subtotal, "subtotal"), (grand_total, "grand total")]
    if tax_rate > 0:
        checks.append((tax_amount, "tax amount"))

    for expected, label in checks:
        # Check both in text (formatted) and in raw numeric values
        formatted = _fmt(expected)
        bare = f"{expected:,.2f}"
        in_text = formatted in text or bare in text
        in_nums = any(abs(n - expected) < 0.02 for n in numbers)
        if in_text or in_nums:
            found += 1
        else:
            issues.append(f"Expected {label} {formatted!r} not found")

    score = round((found / len(checks)) * 10)
    return score, issues


def score_template_fidelity(text: str, t: dict) -> tuple[int, list[str]]:
    issues = []
    found = 0
    checks = [
        (t["business_name"], "business name"),
        (t["quote_data"]["customer_name"], "customer name"),
    ]
    for needle, label in checks:
        if needle and needle.lower() in text.lower():
            found += 1
        else:
            issues.append(f"Not found: {label} '{needle}'")
    score = round((found / len(checks)) * 10)
    return score, issues


def score_field_structure_scratch(text: str, t: dict) -> tuple[int, list[str]]:
    """For scratch XLSX: check expected structural elements are present."""
    issues = []
    found = 0
    checks = [
        ("QUOTATION", "QUOTATION title"),
        ("Bill To", "Bill To section"),
        (t["quote_data"]["customer_address"] or "", "customer address"),
        ("TOTAL", "totals section"),
    ]
    for needle, label in checks:
        if needle and needle.lower() in text.lower():
            found += 1
        else:
            issues.append(f"Structural element not found: {label}")
    score = round((found / len(checks)) * 10)
    return score, issues


def score_field_mapping_template(mapping: dict) -> tuple[int, list[str]]:
    """For template XLSX: score how complete the AI field mapping is."""
    issues = []
    key_fields = [
        ("client_name", "client name cell"),
        ("quote_ref", "quote reference cell"),
        ("quote_date", "quote date cell"),
        ("line_items_start_row", "line items start row"),
        ("total_cell", "total cell"),
    ]
    found = 0
    for key, label in key_fields:
        val = mapping.get(key)
        cols = mapping.get("line_items_cols") or {}
        if key == "line_items_start_row":
            if val and any(cols.values()):
                found += 1
            else:
                issues.append(f"Field not mapped: {label}")
        elif val:
            found += 1
        else:
            issues.append(f"Field not mapped: {label}")

    score = round((found / len(key_fields)) * 10)
    return score, issues


# ─────────────────────────────────────────────────────────────────────────────
# PIPELINE RUNNERS
# ─────────────────────────────────────────────────────────────────────────────

def run_scratch_pipeline(t: dict) -> dict:
    """Test DocumentFactory.generate_xlsx() with synthetic brand_dna."""
    tid = t["id"]
    result = {"id": tid, "config": t, "generated_path": None, "errors": [], "warnings": [], "mode": "scratch"}

    brand_dna = {
        "business_name": t["business_name"],
        "business_address": t["business_address"],
        "contact_details": t["contact_line"],
        "currency": t["currency"],
        "calculation_methods": {"tax_rate": t["tax_rate"]},
        "primary_color_hex": t["primary_color"],
    }
    if t.get("vat_reg"):
        brand_dna["vat_tax_status"] = t["vat_reg"]

    output_filename = f"{tid}_scratch.xlsx"
    try:
        gen = DocumentFactory.generate_xlsx(t["quote_data"], brand_dna, output_filename)
        src = Path(gen["filepath"])
        dst = GENERATED_DIR / f"{tid}_scratch.xlsx"
        shutil.copy2(str(src), str(dst))
        result["generated_path"] = dst
        result["gen_financials"] = {
            "subtotal": gen["subtotal"],
            "tax_amount": gen["tax_amount"],
            "total": gen["total"],
        }
        log.info(f"[{tid}] Scratch XLSX generated: {dst}  subtotal={gen['subtotal']:.2f}")
    except Exception as e:
        result["errors"].append(f"generate_xlsx failed: {e}")
        log.error(f"[{tid}] generate_xlsx error: {traceback.format_exc()}")

    return result


def run_template_pipeline(t: dict) -> dict:
    """Full AI pipeline: create blank XLSX → DNA → mapping → fill template."""
    from ai_service import AIService

    tid = t["id"]
    result = {
        "id": tid, "config": t,
        "blank_path": None, "generated_path": None,
        "brand_dna": None, "field_mapping": None,
        "errors": [], "warnings": [], "mode": "template",
    }

    # Step 1: Create blank XLSX template
    blank_path = TEMPLATES_DIR / f"{tid}_blank.xlsx"
    try:
        wb = create_blank_xlsx(t)
        wb.save(str(blank_path))
        result["blank_path"] = blank_path
        log.info(f"[{tid}] Blank XLSX saved: {blank_path}")
    except Exception as e:
        result["errors"].append(f"Blank template creation failed: {e}")
        log.error(f"[{tid}] Blank creation error: {traceback.format_exc()}")
        return result

    # Step 2: Extract brand DNA
    log.info(f"[{tid}] Extracting brand DNA from XLSX...")
    try:
        brand_dna = AIService.extract_brand_dna_from_xlsx(str(blank_path))
        if not brand_dna:
            result["errors"].append("extract_brand_dna_from_xlsx returned None")
            return result
        brand_dna["currency"] = t["currency"]
        brand_dna["calculation_methods"] = {"tax_rate": t["tax_rate"]}
        brand_dna["preferred_format"] = "xlsx"
        result["brand_dna"] = brand_dna
        log.info(f"[{tid}] DNA: business_name={brand_dna.get('business_name')!r}")
    except Exception as e:
        result["errors"].append(f"Brand DNA extraction failed: {e}")
        log.error(f"[{tid}] DNA error: {traceback.format_exc()}")
        return result

    # Step 3: Build XLSX field mapping
    log.info(f"[{tid}] Building XLSX field mapping (Gemini)...")
    try:
        mapping = AIService.build_xlsx_field_mapping(str(blank_path), brand_dna)
        if not mapping:
            result["errors"].append("build_xlsx_field_mapping returned None")
            return result
        brand_dna["xlsx_field_mapping"] = mapping
        result["field_mapping"] = mapping
        log.info(f"[{tid}] Mapping: {mapping}")
    except Exception as e:
        result["errors"].append(f"Field mapping failed: {e}")
        log.error(f"[{tid}] Mapping error: {traceback.format_exc()}")
        return result

    # Step 4: Fill template
    log.info(f"[{tid}] Filling XLSX template...")
    output_filename = f"{tid}_template.xlsx"
    try:
        template_bytes = blank_path.read_bytes()
        gen = DocumentFactory.generate_from_xlsx_template(
            template_bytes, t["quote_data"], brand_dna, output_filename
        )
        src = Path(gen["filepath"])
        dst = GENERATED_DIR / f"{tid}_template.xlsx"
        shutil.copy2(str(src), str(dst))
        result["generated_path"] = dst
        result["gen_financials"] = {
            "subtotal": gen["subtotal"],
            "tax_amount": gen["tax_amount"],
            "total": gen["total"],
        }
        log.info(f"[{tid}] Template XLSX generated: {dst}  subtotal={gen['subtotal']:.2f}")
    except Exception as e:
        result["errors"].append(f"Template fill failed: {e}")
        log.error(f"[{tid}] Fill error: {traceback.format_exc()}")

    return result


# ─────────────────────────────────────────────────────────────────────────────
# ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────

def analyze_result(result: dict) -> dict:
    t = result["config"]
    sym = _sym(t["currency"])
    tax_rate = t["tax_rate"]
    mode = result.get("mode", "scratch")

    analysis = {
        "id": result["id"],
        "mode": mode,
        "errors": result.get("errors", []),
        "warnings": result.get("warnings", []),
        "scores": {},
        "issues_by_dim": {},
        "field_mapping": result.get("field_mapping"),
    }

    if not result.get("generated_path"):
        analysis["scores"] = {d: 0 for d in ["line_items", "calculations", "fidelity", "structure"]}
        analysis["issues_by_dim"]["pipeline"] = result.get("errors", ["Pipeline did not complete"])
        return analysis

    gen_text = extract_xlsx_text(result["generated_path"])
    gen_numbers = extract_xlsx_numbers(result["generated_path"])

    s1, i1 = score_line_items(gen_text, t["quote_data"])
    s2, i2 = score_calculations(gen_text, gen_numbers, t["quote_data"], tax_rate, sym)
    s3, i3 = score_template_fidelity(gen_text, t)

    if mode == "template":
        mapping = result.get("field_mapping") or {}
        s4, i4 = score_field_mapping_template(mapping)
        struct_dim = "field_mapping"
    else:
        s4, i4 = score_field_structure_scratch(gen_text, t)
        struct_dim = "structure"

    analysis["scores"] = {
        "line_items_integrity": s1,
        "calculation_accuracy": s2,
        "template_fidelity": s3,
        struct_dim: s4,
        "total": s1 + s2 + s3 + s4,
    }
    analysis["issues_by_dim"] = {
        "line_items_integrity": i1,
        "calculation_accuracy": i2,
        "template_fidelity": i3,
        struct_dim: i4,
    }
    return analysis


# ─────────────────────────────────────────────────────────────────────────────
# REPORT
# ─────────────────────────────────────────────────────────────────────────────

def write_report(results: list[dict], analyses: list[dict], run_date: str, mode: str):
    report_path = RESULTS_DIR / "xlsx_test_analysis.md"

    dim_keys_scratch = ["line_items_integrity", "calculation_accuracy", "template_fidelity", "structure"]
    dim_keys_template = ["line_items_integrity", "calculation_accuracy", "template_fidelity", "field_mapping"]
    dim_keys = dim_keys_template if mode == "template" else dim_keys_scratch

    dim_labels = {
        "line_items_integrity": "Line Items",
        "calculation_accuracy": "Calc. Acc.",
        "template_fidelity": "Fidelity",
        "structure": "Structure",
        "field_mapping": "Field Map",
    }

    lines = []
    lines.append(f"# XLSX Test Results — {run_date}\n")
    lines.append(f"**Mode:** `{mode}`  |  **Templates tested:** {len(results)}  |  **Run date:** {run_date}\n")

    lines.append("## Summary\n")
    header = "| Template | Field Style | " + " | ".join(dim_labels[k] for k in dim_keys) + " | **Total /40** | Status |"
    sep = "|---|---|" + "---|" * len(dim_keys) + "---|---|"
    lines.append(header)
    lines.append(sep)

    all_issues = []

    for analysis in analyses:
        tid = analysis["id"]
        scores = analysis.get("scores", {})
        total = scores.get("total", 0)
        fstyle = next((t.get("field_style", "–") for t in TEMPLATES if t["id"] == tid), "–")
        dims = " | ".join(str(scores.get(k, "–")) for k in dim_keys)
        if analysis.get("errors"):
            status = "💥 ERROR"
            total = 0
        elif total >= 28:
            status = "✅ PASS"
        elif total >= 14:
            status = "⚠️ WARN"
        else:
            status = "❌ FAIL"
        lines.append(f"| `{tid}` | `{fstyle}` | {dims} | **{total}** | {status} |")

        for dim, issues in analysis.get("issues_by_dim", {}).items():
            for issue in issues:
                all_issues.append((tid, dim_labels.get(dim, dim), issue))

    lines.append("")
    lines.append("### Scoring Dimensions (40 pts total)\n")
    lines.append("| Dimension | What is measured |")
    lines.append("|---|---|")
    lines.append("| Line Items | All requested line item descriptions appear in the output |")
    lines.append("| Calc. Acc. | Subtotal, tax, and total figures are mathematically correct |")
    lines.append("| Fidelity | Business name and customer name are present |")
    if mode == "template":
        lines.append("| Field Map | AI correctly mapped client_name, quote_ref, quote_date, line_items, total cells |")
    else:
        lines.append("| Structure | QUOTATION title, Bill To section, customer address, totals section present |")
    lines.append("")

    lines.append("## Per-Template Detail\n")

    for result, analysis in zip(results, analyses):
        t = result["config"]
        tid = t["id"]
        scores = analysis.get("scores", {})
        total = scores.get("total", 0)

        lines.append(f"### {tid}")
        lines.append(f"**Business:** {t['business_name']}  |  **Currency:** {t['currency']}  |  **Tax:** {t['tax_rate']}%  |  **Field Style:** `{t.get('field_style', 'brackets')}`\n")

        if analysis.get("errors"):
            lines.append("**Pipeline Errors:**")
            for e in analysis["errors"]:
                lines.append(f"- ❌ {e}")
            lines.append("")
            continue

        lines.append(f"**Total Score: {total}/40**\n")
        lines.append("| Dimension | Score | Issues |")
        lines.append("|---|---|---|")

        for k in dim_keys:
            s = scores.get(k, "–")
            issues = analysis.get("issues_by_dim", {}).get(k, [])
            issue_str = "; ".join(issues[:2]) if issues else "None"
            if len(issues) > 2:
                issue_str += f" (+{len(issues)-2} more)"
            lines.append(f"| {dim_labels.get(k, k)} | {s}/10 | {issue_str} |")

        if mode == "template" and analysis.get("field_mapping"):
            m = analysis["field_mapping"]
            lines.append(f"\n**Mapping:** client_name=`{m.get('client_name')}` | quote_ref=`{m.get('quote_ref')}` | quote_date=`{m.get('quote_date')}` | total=`{m.get('total_cell')}` | li_start=`{m.get('line_items_start_row')}`")

        lines.append("")

    lines.append("## All Issues Found\n")
    if all_issues:
        lines.append("| Template | Dimension | Issue |")
        lines.append("|---|---|---|")
        for tid, dim, issue in all_issues:
            lines.append(f"| `{tid}` | {dim} | {issue} |")
    else:
        lines.append("No issues found.")
    lines.append("")

    lines.append("---")
    lines.append(f"*Generated by `xlsx_test_runner.py` — mode: {mode} — {run_date}*")

    report_path.write_text("\n".join(lines), encoding="utf-8")
    log.info(f"Report written: {report_path}")
    return report_path


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="XLSX quote generation test suite")
    parser.add_argument(
        "--mode",
        choices=["scratch", "template"],
        default="scratch",
        help="scratch: test generate_xlsx() from scratch (no AI); template: full AI pipeline with blank XLSX templates",
    )
    parser.add_argument(
        "--ids",
        nargs="*",
        help="Filter to specific template IDs (e.g. --ids 04 06 09)",
    )
    args, _ = parser.parse_known_args()

    run_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    log.info(f"=== XLSX Test Suite — {run_date} (mode: {args.mode}) ===")

    templates_to_run = TEMPLATES
    if args.ids:
        templates_to_run = [t for t in TEMPLATES if any(t["id"].startswith(p) for p in args.ids)]
        if not templates_to_run:
            log.error(f"No templates matched IDs: {args.ids}")
            sys.exit(1)

    log.info(f"Running {len(templates_to_run)} template(s)...")

    results = []
    analyses = []

    for i, t in enumerate(templates_to_run, 1):
        log.info(f"\n{'='*60}")
        log.info(f"[{i}/{len(templates_to_run)}] {t['id']} ({t['business_name']})")
        log.info(f"{'='*60}")
        if args.mode == "template":
            result = run_template_pipeline(t)
        else:
            result = run_scratch_pipeline(t)
        results.append(result)

    log.info(f"\n{'='*60}")
    log.info("Analysing results...")
    log.info(f"{'='*60}\n")

    for result in results:
        log.info(f"Analysing: {result['id']}...")
        analysis = analyze_result(result)
        analyses.append(analysis)
        total = analysis.get("scores", {}).get("total", "n/a")
        errors = len(analysis.get("errors", []))
        log.info(f"  → Score: {total}/40  |  Errors: {errors}")

    report_path = write_report(results, analyses, run_date, args.mode)

    print(f"\n{'='*60}")
    print(f"XLSX TEST RESULTS  (mode: {args.mode})")
    print(f"{'='*60}")
    print(f"{'Template':<25} {'Score':>7}  Status")
    print(f"{'-'*45}")
    for analysis in analyses:
        tid = analysis["id"]
        total = analysis.get("scores", {}).get("total", 0)
        if analysis.get("errors"):
            status = "ERROR"
        elif total >= 28:
            status = "PASS"
        elif total >= 14:
            status = "WARN"
        else:
            status = "FAIL"
        print(f"{tid:<25} {total:>5}/40  {status}")
    print(f"\nReport: {report_path.absolute()}")


if __name__ == "__main__":
    main()
